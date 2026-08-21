"""Contract tests for the approved DEC Kobo XLSForms (no network or secrets)."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
QUICK = ROOT / "forms" / "DEC_Internal_Pilot_Quick_Finding_XLSForm_v2_1_20260819.xlsx"
REVIEW = ROOT / "forms" / "DEC_Internal_Pilot_Final_Course_Review_XLSForm_v2_1_20260819.xlsx"


def rows(path: Path, sheet: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    values = list(workbook[sheet].iter_rows(values_only=True))
    headers = [str(value or "") for value in values[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index] and value is not None}
        for row in values[1:]
        if any(value is not None for value in row)
    ]


def by_name(path: Path) -> dict[str, dict[str, object]]:
    return {str(row["name"]): row for row in rows(path, "survey") if row.get("name")}


def choices(path: Path, list_name: str) -> list[str]:
    return [str(row["name"]) for row in rows(path, "choices") if row.get("list_name") == list_name]


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


manifest = json.loads((ROOT / "SOURCE_MANIFEST.json").read_text(encoding="utf-8"))
manifest_hashes = {entry["path"]: entry["sha256"] for entry in manifest["files"]}
assert manifest["current_authoritative_form_version"] == "20260819_v2_1"
assert sha256(QUICK) == manifest_hashes[str(QUICK.relative_to(ROOT)).replace("\\", "/")]
assert sha256(REVIEW) == manifest_hashes[str(REVIEW.relative_to(ROOT)).replace("\\", "/")]

quick = by_name(QUICK)
quick_settings = rows(QUICK, "settings")[0]
assert quick_settings["form_id"] == "dec_internal_pilot_quick_finding"
assert quick_settings["version"] == "20260819_v2_1"
assert set(quick) == {"start", "end", "qf_intro", "tester_id", "identity_missing", "observation_location", "stable_id", "what_happened", "recommendation", "screenshot", "urgent_note"}
assert quick["tester_id"]["type"] == "hidden" and quick["tester_id"]["required"] == "yes"
assert quick["screenshot"]["type"] == "image"
assert choices(QUICK, "observation_location") == ["hub", "hrba", "pm"]
for removed in ("issue_device_browser", "finding_category", "finding_category_other", "blocker_flag", "blocker_note"):
    assert removed not in quick

review = by_name(REVIEW)
review_settings = rows(REVIEW, "settings")[0]
assert review_settings["form_id"] == "dec_internal_pilot_final_review_readiness"
assert review_settings["version"] == "20260819_v2_1"
assert review["tester_id"]["type"] == "hidden" and review["tester_id"]["required"] == "yes"
assert review["course"]["type"] == "hidden" and review["course"]["required"] == "yes"
assert review["review_key"]["calculation"] == "concat(${tester_id}, '-', ${course})"
assert choices(REVIEW, "course") == ["hrba", "pm"]
assert choices(REVIEW, "practical_result") == ["pass", "pass_issue", "fail", "not_tested"]
assert choices(REVIEW, "quality_rating") == ["blocked_0", "fragile_1", "workable_2", "strong_3", "not_tested"]

practical = [f"b{index:02d}_{suffix}" for index, suffix in enumerate([
    "account_activation", "sign_in", "course_access", "course_separation", "start_resume", "required_gating", "correct_progression", "progress_persistence", "device_use", "media", "final_assessment", "completion_certificate", "return_completed", "cross_user", "feedback_support", "second_device"
], start=1)]
assert all(name in review for name in practical)
assert "not_tested" in str(review["b_critical_not_tested_reason"]["relevant"])
assert review["sec_e_hrba"]["relevant"] == "${course} = 'hrba'"
assert review["sec_e_pm"]["relevant"] == "${course} = 'pm'"

quality_base = [name for name, row in review.items() if str(row.get("type", "")).startswith("select_one quality_rating")]
assert len(quality_base) == 43
for name in quality_base:
    comment = review[f"{name}_comment"]
    relevant = str(comment["relevant"])
    assert relevant == f"${{{name}}} = 'blocked_0' or ${{{name}}} = 'fragile_1'"
    assert "workable_2" not in relevant

for open_text in ("c_access_improvement", "d_difficult_screen", "e_best_decision_activity", "f_realistic_relevant", "f_unrealistic_unclear", "g_practical_example", "g_support_needs", "h_staff_support_improvement", "j_keep", "j_priority_improvement", "j_blocker_explain", "l_readiness_reason"):
    assert open_text in review

print(json.dumps({"status": "pass", "version": "20260819_v2_1", "quick_named_rows": len(quick), "review_named_rows": len(review), "quality_comment_rules": len(quality_base)}))
